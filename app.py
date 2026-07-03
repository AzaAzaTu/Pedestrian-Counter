from pathlib import Path
from uuid import uuid4
from datetime import datetime

from flask import Flask, render_template, request, url_for, flash, send_file
from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from utils.detector import PedestrianDetector
from utils.db import init_db, add_request, get_history


BASE_DIR = Path(__file__).resolve().parent

UPLOAD_DIR = BASE_DIR / "static" / "uploads"
RESULT_DIR = BASE_DIR / "static" / "results"
REPORTS_DIR = BASE_DIR / "reports"

ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "webp"}


app = Flask(__name__)
app.secret_key = "pedestrian-counter-secret-key"

app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
RESULT_DIR.mkdir(parents=True, exist_ok=True)
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

init_db()

detector = PedestrianDetector(
    model_path="yolov8n.pt",
    confidence=0.25
)


def allowed_file(filename):
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS
    )


def get_pdf_font():
    font_paths = [
        BASE_DIR / "static" / "fonts" / "DejaVuSans.ttf",
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
    ]

    for font_path in font_paths:
        if font_path.exists():
            pdfmetrics.registerFont(TTFont("AppFont", str(font_path)))
            return "AppFont"

    return "Helvetica"


@app.route("/", methods=["GET", "POST"])
def index():
    result_image = None
    uploaded_image = None
    pedestrian_count = None
    objects = []

    if request.method == "POST":
        if "image" not in request.files:
            flash("Файл не был передан.")
            return render_template("index.html", history=get_history())

        file = request.files["image"]

        if file.filename == "":
            flash("Выберите изображение.")
            return render_template("index.html", history=get_history())

        if not allowed_file(file.filename):
            flash("Разрешены только файлы JPG, JPEG, PNG и WEBP.")
            return render_template("index.html", history=get_history())

        original_filename = secure_filename(file.filename)
        extension = original_filename.rsplit(".", 1)[1].lower()

        unique_name = f"{uuid4().hex}.{extension}"
        upload_path = UPLOAD_DIR / unique_name

        result_name = f"result_{uuid4().hex}.jpg"
        result_path = RESULT_DIR / result_name

        file.save(upload_path)

        try:
            detection_result = detector.detect(upload_path, result_path)

            pedestrian_count = detection_result["count"]
            objects = detection_result["objects"]

            uploaded_image = url_for("static", filename=f"uploads/{unique_name}")
            result_image = url_for("static", filename=f"results/{result_name}")

            add_request(
                original_filename=original_filename,
                upload_path=f"uploads/{unique_name}",
                result_path=f"results/{result_name}",
                pedestrian_count=pedestrian_count,
                confidence=detector.confidence
            )

        except Exception as error:
            flash(f"Ошибка обработки изображения: {error}")

    history = get_history()

    return render_template(
        "index.html",
        uploaded_image=uploaded_image,
        result_image=result_image,
        pedestrian_count=pedestrian_count,
        objects=objects,
        history=history
    )


@app.route("/export/excel")
def export_excel():
    history = get_history(limit=1000)

    wb = Workbook()
    ws = wb.active
    ws.title = "История обработки"

    ws["A1"] = "История обработки изображений"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    headers = [
        "№",
        "Дата и время",
        "Файл",
        "Количество пешеходов",
        "Порог уверенности"
    ]

    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for index, row in enumerate(history, start=1):
        ws.append([
            index,
            row["timestamp"],
            row["original_filename"],
            row["pedestrian_count"],
            row["confidence"]
        ])

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 20

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    file_path = REPORTS_DIR / "pedestrian_history.xlsx"
    wb.save(file_path)

    return send_file(
        file_path,
        as_attachment=True,
        download_name="pedestrian_history.xlsx"
    )


@app.route("/export/pdf")
def export_pdf():
    history = get_history(limit=1000)

    file_path = REPORTS_DIR / "pedestrian_history.pdf"

    font_name = get_pdf_font()

    doc = SimpleDocTemplate(
        str(file_path),
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    title_style = ParagraphStyle(
        name="Title",
        fontName=font_name,
        fontSize=16,
        leading=20,
        alignment=1,
        spaceAfter=20
    )

    text_style = ParagraphStyle(
        name="Text",
        fontName=font_name,
        fontSize=10,
        leading=12
    )

    elements = []

    elements.append(Paragraph("История обработки изображений", title_style))
    elements.append(Paragraph("Проект: Подсчет пешеходов на переходе", text_style))
    elements.append(Paragraph("Приложение: Pedestrian Counter v 1.0.0", text_style))
    elements.append(Paragraph("Автор: AzaTu", text_style))
    elements.append(Paragraph(f"Дата формирования: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", text_style))
    elements.append(Spacer(1, 15))

    data = [
        [
            Paragraph("№", text_style),
            Paragraph("Дата и время", text_style),
            Paragraph("Файл", text_style),
            Paragraph("Пешеходы", text_style),
            Paragraph("Уверенность", text_style)
        ]
    ]

    for index, row in enumerate(history, start=1):
        data.append([
            Paragraph(str(index), text_style),
            Paragraph(str(row["timestamp"]), text_style),
            Paragraph(str(row["original_filename"]), text_style),
            Paragraph(str(row["pedestrian_count"]), text_style),
            Paragraph(str(row["confidence"]), text_style)
        ])

    if not history:
        data.append([
            Paragraph("-", text_style),
            Paragraph("История пуста", text_style),
            Paragraph("-", text_style),
            Paragraph("-", text_style),
            Paragraph("-", text_style)
        ])

    table = Table(
        data,
        colWidths=[30, 105, 190, 70, 80]
    )

    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (3, 0), (4, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, -1), font_name)
    ]))

    elements.append(table)

    doc.build(elements)

    return send_file(
        file_path,
        as_attachment=True,
        download_name="pedestrian_history.pdf"
    )


if __name__ == "__main__":
    app.run(debug=True)
